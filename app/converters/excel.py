import io
import time
from typing import List, Any, Optional
from loguru import logger

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.report import ReportData


def csv_to_excel_buffer(report_data: ReportData) -> io.BytesIO:
    """
    Конвертация данных отчета в Excel файл в памяти

    Args:
        report_data: Данные отчета с заголовками и строками

    Returns:
        BytesIO объект с Excel файлом
    """
    try:
        # Создаем новую рабочую книгу
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Отчет"

        # Записываем заголовки
        if report_data.headers:
            for col_idx, header in enumerate(report_data.headers, start=1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                _apply_header_style(cell)

        # Записываем данные
        if report_data.rows:
            for row_idx, row_data in enumerate(report_data.rows, start=2):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    _apply_data_style(cell, col_idx)

        # Автоподбор ширины колонок
        _adjust_column_widths(worksheet)

        # Добавляем мета-данные если есть
        if report_data.meta_data:
            _add_metadata_sheet(workbook, report_data.meta_data)

        # Сохраняем в BytesIO
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info(f"✅ Successfully converted report to Excel format")
        return excel_buffer

    except Exception as e:
        logger.error(f"❌ Error converting report to Excel: {str(e)}")
        raise


def _apply_header_style(cell) -> None:
    """Применение стиля к ячейке заголовка"""
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _apply_data_style(cell, col_idx: int) -> None:
    """Применение стиля к ячейке с данными"""
    cell.font = Font(name="Calibri", size=10)

    # Чередующиеся цвета строк
    if cell.row % 2 == 0:
        cell.fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

    # Выравнивание в зависимости от типа данных
    if isinstance(cell.value, (int, float)):
        cell.alignment = Alignment(horizontal="right", vertical="center")
        # Форматирование чисел
        if isinstance(cell.value, float):
            cell.number_format = "#,##0.00"
        else:
            cell.number_format = "#,##0"
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    cell.border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )


def _adjust_column_widths(worksheet: Worksheet) -> None:
    """Автоподбор ширины колонок"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Устанавливаем ширину с ограничениями
        adjusted_width = min(max(max_length + 2, 8), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _add_metadata_sheet(workbook: Workbook, meta_data: List[Any]) -> None:
    """Добавление листа с метаданными"""
    if not meta_data:
        return

    try:
        meta_sheet = workbook.create_sheet("Метаданные")

        for row_idx, meta_item in enumerate(meta_data, start=1):
            if isinstance(meta_item, list):
                for col_idx, value in enumerate(meta_item, start=1):
                    meta_sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                meta_sheet.cell(row=row_idx, column=1, value=str(meta_item))

        # Применяем базовые стили
        for row in meta_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        _adjust_column_widths(meta_sheet)

    except Exception as e:
        logger.warning(f"Failed to add metadata sheet: {str(e)}")


def generate_excel_filename(prefix: str = "report") -> str:
    """
    Генерация имени Excel файла

    Args:
        prefix: Префикс имени файла

    Returns:
        Имя файла с расширением .xlsx
    """
    timestamp = int(time.time())
    return f"{prefix}_{timestamp}.xlsx"


def validate_excel_data(report_data: ReportData) -> bool:
    """
    Валидация данных перед конвертацией в Excel

    Args:
        report_data: Данные отчета

    Returns:
        True если данные валидны
    """
    if not report_data:
        logger.error("Report data is None")
        return False

    if not report_data.headers:
        logger.warning("Report has no headers")
        return False

    if not report_data.rows:
        logger.warning("Report has no data rows")
        return False

    # Проверяем соответствие количества колонок
    header_count = len(report_data.headers)
    for idx, row in enumerate(report_data.rows):
        if len(row) != header_count:
            logger.warning(f"Row {idx} has {len(row)} columns, expected {header_count}")
            # Дополняем или обрезаем строку до нужной длины
            if len(row) < header_count:
                report_data.rows[idx].extend([None] * (header_count - len(row)))
            else:
                report_data.rows[idx] = row[:header_count]

    return True


def test_excel_conversion() -> bool:
    """
    Тестовая функция для проверки работы Excel конвертера

    Returns:
        True если тест прошел успешно
    """
    try:
        # Создаем тестовые данные
        test_data = ReportData(
            headers=["Дата", "Расходы", "Клики", "CPC", "Визиты"],
            rows=[
                ["2025-01-01", 1000.50, 250, 4.00, 180],
                ["2025-01-02", 1200.75, 300, 4.01, 210],
                ["2025-01-03", 950.25, 190, 5.00, 160],
            ],
            meta_data=["Тестовый отчет", "Создан автоматически"],
        )

        # Валидируем данные
        if not validate_excel_data(test_data):
            logger.error("Test data validation failed")
            return False

        # Конвертируем в Excel
        excel_buffer = csv_to_excel_buffer(test_data)

        # Проверяем что буфер не пустой
        if excel_buffer.getbuffer().nbytes == 0:
            logger.error("Excel buffer is empty")
            return False

        logger.info(
            f"✅ Excel conversion test passed. File size: {excel_buffer.getbuffer().nbytes} bytes"
        )
        return True

    except Exception as e:
        logger.error(f"❌ Excel conversion test failed: {str(e)}")
        return False


def read_excel_preview(excel_content: bytes, max_rows: int = 20) -> Optional[dict]:
    """
    Чтение первых строк из Excel файла для превью

    Args:
        excel_content: Содержимое Excel файла в байтах
        max_rows: Максимальное количество строк для превью

    Returns:
        Словарь с headers, rows и total_rows или None при ошибке
    """
    try:
        from openpyxl import load_workbook

        # Загружаем Excel файл из байтов
        excel_buffer = io.BytesIO(excel_content)
        workbook = load_workbook(excel_buffer, read_only=True)

        # Берем первый лист
        worksheet = workbook.active

        # Получаем все строки
        all_rows = list(worksheet.iter_rows(values_only=True))

        if not all_rows:
            logger.warning("Excel file is empty")
            return None

        # Первая строка - заголовки
        headers = [str(cell) if cell is not None else "" for cell in all_rows[0]]

        # Данные (пропускаем заголовок и берем первые max_rows строк)
        data_rows = all_rows[1 : max_rows + 1]

        # Преобразуем данные в нужный формат
        preview_rows = []
        for row in data_rows:
            formatted_row = []
            for cell in row:
                if cell is None:
                    formatted_row.append("")
                elif isinstance(cell, (int, float)):
                    formatted_row.append(cell)
                else:
                    formatted_row.append(str(cell))
            preview_rows.append(formatted_row)

        # Общее количество строк (включая заголовок)
        total_rows = len(all_rows) - 1  # Не считаем заголовок

        workbook.close()

        result = {"headers": headers, "rows": preview_rows, "total_rows": total_rows}

        logger.info(
            f"✅ Excel preview generated: {len(headers)} columns, {len(preview_rows)} preview rows, {total_rows} total rows"
        )
        return result

    except Exception as e:
        logger.error(f"❌ Error reading Excel preview: {str(e)}")
        return None


async def get_report_data_preview(file_name: str) -> Optional[dict]:
    """
    Получение превью данных отчета из S3

    Args:
        file_name: Имя файла в S3

    Returns:
        Словарь с превью данных или None при ошибке
    """
    try:
        from app.services.s3_service import s3_service

        # Скачиваем файл из S3
        excel_content = await s3_service.download_report_content(file_name)
        if not excel_content:
            logger.error(f"Failed to download file: {file_name}")
            return None

        # Читаем превью из Excel файла
        preview_data = read_excel_preview(excel_content)
        if not preview_data:
            logger.error(f"Failed to read Excel preview: {file_name}")
            return None

        return preview_data

    except Exception as e:
        logger.error(f"❌ Error getting report data preview: {str(e)}")
        return None
